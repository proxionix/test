#!/usr/bin/env python3
import os
import re
import imaplib
import email
from email.header import decode_header
from flask import Flask, render_template, request, redirect, url_for, flash, send_from_directory, jsonify
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell  # Ajout de l'import pour MergedCell
from flask_mail import Mail, Message
import logging
import subprocess
import datetime
import json
import random
from flask_socketio import SocketIO
import shutil
from datetime import datetime, timedelta
import threading
import time
import schedule

app = Flask(__name__, template_folder="templates", static_folder="static")
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'supersecretkey')
socketio = SocketIO(app)

# Configuration de Flask-Mail depuis les variables d'environnement
app.config.update(
    MAIL_SERVER=os.environ.get('MAIL_SERVER', 'smtp.example.com'),
    MAIL_PORT=int(os.environ.get('MAIL_PORT', '587')),
    MAIL_USE_TLS=os.environ.get('MAIL_USE_TLS', 'True').lower() in ['true', '1', 'yes'],
    MAIL_USERNAME=os.environ.get('MAIL_USERNAME', 'your_email@example.com'),
    MAIL_PASSWORD=os.environ.get('MAIL_PASSWORD', 'your_email_password'),
    MAIL_DEFAULT_SENDER=os.environ.get('MAIL_DEFAULT_SENDER', 'your_email@example.com')
)
mail = Mail(app)

# Configuration IMAP pour iCloud
IMAP_SERVER = "imap.mail.me.com"
IMAP_PORT = 993

# Répertoire de stockage
DOWNLOAD_DIR = "telechargements_xls"  # Même nom que dans votre script initial
REFUSED_DIR = "refused"
ARCHIVE_DIR = "archives"  # Nouveau répertoire pour les archives

# Chemin vers LibreOffice
LIBREOFFICE_PATH = "C:\\Program Files\\LibreOffice\\program\\soffice.exe"

# Création des répertoires s'ils n'existent pas
if not os.path.exists(DOWNLOAD_DIR):
    os.makedirs(DOWNLOAD_DIR)
if not os.path.exists(REFUSED_DIR):
    os.makedirs(REFUSED_DIR)
if not os.path.exists(ARCHIVE_DIR):
    os.makedirs(ARCHIVE_DIR)
if not os.path.exists('static/uploads'):
    os.makedirs('static/uploads')

# Stockage des tickets en mémoire (pour une solution production, utiliser une BDD)
tickets = {}
notification_count = 0

# Configuration du logging pour faciliter le debug
logging.basicConfig(level=logging.DEBUG)

def decode_mime_words(s):
    """Décode les mots encodés en MIME dans les en-têtes."""
    if not s:
        return ""
    try:
        decoded_words = decode_header(s)
        return ''.join([
            part[0].decode(part[1] if part[1] else "utf-8") if isinstance(part[0], bytes) else str(part[0])
            for part in decoded_words
        ])
    except Exception as e:
        app.logger.error(f"Erreur lors du décodage MIME: {e}")
        return str(s)

def extract_sr_folder_name(subject):
    """
    Extrait le nom du dossier basé sur la partie après 'SR-' dans l'objet.
    Transforme le format pour obtenir SR-Chiffre-Texte sans caractères spéciaux.
    """
    if not subject:
        return None
        
    sr_match = re.search(r'SR-([^\r\n]+)', subject)
    if sr_match:
        folder_content = sr_match.group(1).strip()
        digits_match = re.search(r'(\d+)', folder_content)
        digits_part = digits_match.group(1) if digits_match else ""
        text_part = ""
        if digits_match:
            end_pos = digits_match.end()
            if end_pos < len(folder_content):
                text_part = folder_content[end_pos:].strip()
                text_part = re.sub(r'[^a-zA-Z0-9]', '-', text_part)
                text_part = re.sub(r'-+', '-', text_part)
                text_part = text_part.strip('-')
        if text_part:
            folder_name = f"SR-{digits_part}-{text_part}"
        else:
            folder_name = f"SR-{digits_part}"
        return folder_name
    return None

def sanitize_filename(filename):
    """Assure que le nom de fichier est valide pour le système de fichiers."""
    if not filename:
        return "unknown_file"
    return re.sub(r'["\n\r\t\\/:*?<>|]', '_', filename)

def determine_client(subject, sender):
    """
    Détermine le client en fonction du sujet ou de l'expéditeur.
    """
    subject_lower = subject.lower() if subject else ""
    sender_lower = sender.lower() if sender else ""
    
    if "infrabel" in subject_lower or "infrabel" in sender_lower:
        return "Infrabel"
    else:
        return "Bnp Paribas Fortis"  # Client par défaut

def determine_priority(subject):
    """
    Détermine la priorité du ticket en fonction du contenu du sujet.
    """
    subject_lower = subject.lower() if subject else ""
    
    if "urgent" in subject_lower or "urgence" in subject_lower or "critique" in subject_lower:
        return "high"
    elif "important" in subject_lower or "prioritaire" in subject_lower:
        return "medium"
    else:
        return "normal"

def fetch_emails():
    """
    Récupère les emails et stocke les informations en mémoire,
    mais ne télécharge pas encore les pièces jointes.
    """
    global tickets, notification_count
    user_email = os.environ.get('ICLOUD_USER_EMAIL')
    user_password = os.environ.get('ICLOUD_USER_PASSWORD')
    if not user_email or not user_password:
        msg = ("Configuration de l'email introuvable. "
               "Veuillez définir ICLOUD_USER_EMAIL et ICLOUD_USER_PASSWORD.")
        app.logger.error(msg)
        flash(msg, "danger")
        return

    try:
        mail_conn = imaplib.IMAP4_SSL(IMAP_SERVER, IMAP_PORT)
        mail_conn.login(user_email, user_password)
    except Exception as e:
        app.logger.error(f"Erreur lors de la connexion ou de l'authentification : {e}")
        flash(f"Erreur lors de la connexion ou de l'authentification : {e}", "danger")
        return

    status, _ = mail_conn.select("INBOX")
    if status != "OK":
        flash("Erreur lors de la sélection de la boîte INBOX.", "danger")
        return

    # Utilisation de 'ALL' pour reproduire le comportement du script initial
    status, message_numbers = mail_conn.search(None, 'ALL')
    if status != "OK":
        flash("Erreur lors de la recherche des emails.", "danger")
        return

    if not message_numbers or not message_numbers[0]:
        app.logger.info("Aucun email trouvé.")
        mail_conn.logout()
        return

    message_ids = message_numbers[0].split()

    app.logger.debug(f"{len(message_ids)} email(s) trouvé(s).")
    
    # Nombre initial de tickets
    initial_ticket_count = len(tickets)

    for msg_id in message_ids:
        msg_id_str = msg_id.decode() if isinstance(msg_id, bytes) else str(msg_id)
        
        # Si le ticket existe déjà, passer au suivant
        if msg_id_str in tickets:
            continue
            
        try:
            # Ne pas télécharger le contenu de l'email, juste les informations de base
            status, msg_data = mail_conn.fetch(msg_id, "(RFC822.HEADER)")
            if status != "OK":
                app.logger.warning(f"Erreur lors de la récupération de l'en-tête du message {msg_id_str}.")
                continue

            raw_email_header = None
            for item in msg_data:
                if isinstance(item, tuple) and len(item) > 1:
                    raw_email_header = item[1]
                    break

            if raw_email_header is None:
                app.logger.warning(f"Aucun en-tête email trouvé pour le message {msg_id_str}.")
                continue

            msg = email.message_from_bytes(raw_email_header)
            subject = decode_mime_words(msg.get("Subject", "Aucun sujet"))
            sender = decode_mime_words(msg.get("From", "Inconnu"))
            sr_folder_name = extract_sr_folder_name(subject)
            
            # Récupérer la date de l'email
            date_str = msg.get("Date")
            email_date = None
            if date_str:
                try:
                    # Convertir la date en objet datetime
                    from email.utils import parsedate_to_datetime
                    email_date = parsedate_to_datetime(date_str)
                    email_date = email_date.strftime("%Y-%m-%d %H:%M")
                except Exception as e:
                    app.logger.warning(f"Erreur lors de la conversion de la date: {e}")
                    email_date = None
            
            # Déterminer le client et la priorité
            client = determine_client(subject, sender)
            priority = determine_priority(subject)

            # Ajout d'un identifiant unique basé sur la date et un nombre aléatoire
            ticket_id = f"T{datetime.now().strftime('%Y%m%d')}-{random.randint(1000, 9999)}"

            # Ajout d'un attribut client pour différencier les tickets
            tickets[msg_id_str] = {
                'subject': subject,
                'sender': sender,
                'status': 'new',
                'filepath': None,
                'original_email_id': msg_id_str,
                'ticket_id': ticket_id,
                'sr_folder_name': sr_folder_name,
                'client': client,
                'priority': priority,
                'intervention_details': '',
                'date_reception': email_date or datetime.now().strftime("%Y-%m-%d %H:%M"),
                'date_traitement': None,
                'date_envoi': None,
                'tags': []
            }
        except Exception as e:
            app.logger.error(f"Erreur lors du traitement du message {msg_id_str} : {e}")
            continue

    mail_conn.logout()
    
    # Si de nouveaux tickets ont été ajoutés, augmenter le compteur de notifications
    new_tickets_count = len(tickets) - initial_ticket_count
    if new_tickets_count > 0:
        notification_count += new_tickets_count
        socketio.emit('new_tickets', {'count': new_tickets_count, 'total': notification_count})
        flash(f"{new_tickets_count} nouveau(x) ticket(s) reçu(s).", "success")
    
    # Sauvegarder l'état des tickets
    save_tickets_state()

def download_email_attachment(msg_id_str):
    """
    Télécharge les pièces jointes d'un email spécifique.
    Retourne le chemin du fichier téléchargé.
    """
    user_email = os.environ.get('ICLOUD_USER_EMAIL')
    user_password = os.environ.get('ICLOUD_USER_PASSWORD')
    if not user_email or not user_password:
        raise Exception("Configuration de l'email introuvable.")

    mail_conn = imaplib.IMAP4_SSL(IMAP_SERVER, IMAP_PORT)
    mail_conn.login(user_email, user_password)
    mail_conn.select("INBOX")

    # Récupérer l'email complet
    status, msg_data = mail_conn.fetch(msg_id_str.encode() if isinstance(msg_id_str, str) else msg_id_str, "(BODY[])")
    if status != "OK":
        mail_conn.logout()
        raise Exception(f"Erreur lors de la récupération du message {msg_id_str}.")

    raw_email = None
    for item in msg_data:
        if isinstance(item, tuple) and len(item) > 1:
            raw_email = item[1]
            break

    if raw_email is None:
        mail_conn.logout()
        raise Exception(f"Aucun contenu email trouvé pour le message {msg_id_str}.")

    msg = email.message_from_bytes(raw_email)
    sr_folder_name = tickets[msg_id_str]['sr_folder_name']
    
    # Créer le dossier si nécessaire
    if sr_folder_name:
        subfolder_path = os.path.join(DOWNLOAD_DIR, sr_folder_name)
        if not os.path.exists(subfolder_path):
            os.makedirs(subfolder_path)
    
    file_path = None
    
    # Parcourir les pièces jointes
    for part in msg.walk():
        filename = part.get_filename()
        if filename:
            filename = decode_mime_words(filename)
            filename = sanitize_filename(filename)
            if filename.lower().endswith(".xls") or filename.lower().endswith(".xlsx"):
                # Télécharger la pièce jointe
                if sr_folder_name:
                    file_path = os.path.join(subfolder_path, filename)
                else:
                    file_path = os.path.join(DOWNLOAD_DIR, filename)
                
                with open(file_path, 'wb') as f:
                    f.write(part.get_payload(decode=True))
                
                app.logger.info(f"Fichier téléchargé: {file_path}")
                break  # On suppose qu'il n'y a qu'une pièce jointe Excel qui nous intéresse
    
    mail_conn.logout()
    return file_path

def convert_xls_to_xlsx(file_path):
    """
    Convertit un fichier .xls en .xlsx en utilisant LibreOffice.
    Supprime le fichier XLS original après une conversion réussie.
    Retourne le chemin du nouveau fichier si la conversion réussit, sinon None.
    """
    if not file_path or not file_path.lower().endswith('.xls'):
        return file_path
    
    app.logger.info(f"Conversion du fichier {file_path} de XLS à XLSX en utilisant LibreOffice...")
    
    # Dossier de sortie (même que le dossier d'entrée)
    output_dir = os.path.dirname(file_path)
    
    try:
        # Chemin absolu du fichier d'entrée
        abs_file_path = os.path.abspath(file_path)
        
        # Vérifier que LibreOffice existe
        if not os.path.exists(LIBREOFFICE_PATH):
            app.logger.error(f"LibreOffice introuvable à {LIBREOFFICE_PATH}")
            return None
        
        # Convertir avec LibreOffice
        subprocess.run([
            LIBREOFFICE_PATH,
            "--headless",
            "--convert-to", "xlsx",
            "--outdir", output_dir,
            abs_file_path
        ], check=True)
        
        # Le nouveau fichier a le même nom mais avec l'extension .xlsx
        base_name = os.path.basename(file_path)
        base_name_without_ext = os.path.splitext(base_name)[0]
        new_file_path = os.path.join(output_dir, base_name_without_ext + ".xlsx")
        
        if os.path.exists(new_file_path):
            app.logger.info(f"Conversion réussie. Nouveau fichier: {new_file_path}")
            
            # Supprimer le fichier XLS original
            try:
                os.remove(file_path)
                app.logger.info(f"Fichier XLS original supprimé: {file_path}")
            except Exception as rm_err:
                app.logger.warning(f"Impossible de supprimer le fichier XLS original: {rm_err}")
            
            return new_file_path
        else:
            app.logger.error("Échec de la conversion: fichier de sortie introuvable.")
            return None
    
    except Exception as e:
        app.logger.error(f"Erreur lors de la conversion: {e}")
        return None

# Fonction pour supprimer les fichiers plus vieux que 7 jours
def cleanup_old_files():
    """
    Supprime les fichiers traités qui ont plus de 7 jours et marque les tickets correspondants comme 'archived'
    """
    app.logger.info("Exécution du nettoyage automatique des fichiers...")
    now = datetime.now()
    tickets_to_archive = []
    seven_days_ago = now - timedelta(days=7)
    
    # Parcourir tous les tickets qui ont été traités et envoyés
    for ticket_id, ticket in tickets.items():
        if ticket['status'] == 'sent' and ticket.get('date_envoi'):
            # Convertir la date d'envoi en objet datetime
            try:
                date_envoi = datetime.strptime(ticket['date_envoi'], "%Y-%m-%d %H:%M")
                
                # Vérifier si le ticket a plus de 7 jours
                if date_envoi < seven_days_ago:
                    # Ajouter le ticket à la liste pour archivage
                    tickets_to_archive.append(ticket_id)
                    
                    # Créer un dossier d'archive si nécessaire
                    archive_date_dir = os.path.join(ARCHIVE_DIR, date_envoi.strftime("%Y-%m"))
                    if not os.path.exists(archive_date_dir):
                        os.makedirs(archive_date_dir)
                    
                    # Déplacer le fichier associé s'il existe vers les archives
                    if ticket.get('filepath') and os.path.exists(ticket['filepath']):
                        try:
                            file_name = os.path.basename(ticket['filepath'])
                            archive_path = os.path.join(archive_date_dir, file_name)
                            shutil.move(ticket['filepath'], archive_path)
                            ticket['filepath'] = archive_path  # Mettre à jour le chemin
                            app.logger.info(f"Fichier archivé: {archive_path}")
                        except Exception as e:
                            app.logger.error(f"Erreur lors de l'archivage du fichier {ticket['filepath']}: {e}")
                    
                    # Si un dossier SR a été créé et qu'il est vide, le supprimer
                    if ticket.get('sr_folder_name'):
                        sr_folder_path = os.path.join(DOWNLOAD_DIR, ticket['sr_folder_name'])
                        if os.path.exists(sr_folder_path) and not os.listdir(sr_folder_path):
                            try:
                                shutil.rmtree(sr_folder_path)
                                app.logger.info(f"Dossier vide supprimé: {sr_folder_path}")
                            except Exception as e:
                                app.logger.error(f"Erreur lors de la suppression du dossier {sr_folder_path}: {e}")
            except Exception as e:
                app.logger.error(f"Erreur lors du traitement de la date d'envoi pour ticket {ticket_id}: {e}")
    
    # Marquer les tickets comme archivés
    for ticket_id in tickets_to_archive:
        tickets[ticket_id]['status'] = 'archived'
        app.logger.info(f"Ticket {ticket_id} archivé après 7 jours")
    
    # Émettre une notification pour informer de l'archivage
    if tickets_to_archive:
        socketio.emit('tickets_archived', {'count': len(tickets_to_archive)})
        save_tickets_state()  # Sauvegarder l'état après archivage
    
    app.logger.info(f"Nettoyage terminé: {len(tickets_to_archive)} tickets archivés")

# Fonction pour le thread de vérification périodique des emails
def scheduled_tasks():
    """
    Exécute les tâches planifiées en arrière-plan:
    - Vérification des emails à intervalles réguliers
    - Nettoyage des fichiers anciens
    """
    schedule.every(5).minutes.do(fetch_emails)
    schedule.every().day.at("01:00").do(cleanup_old_files)  # Nettoyage à 1h du matin
    
    while True:
        schedule.run_pending()
        time.sleep(60)  # Vérifier toutes les minutes

# Fonction pour démarrer le thread des tâches planifiées
def start_scheduled_tasks():
    """Démarre le thread pour les tâches planifiées"""
    thread = threading.Thread(target=scheduled_tasks, daemon=True)
    thread.start()
    app.logger.info("Thread de tâches planifiées démarré")

# Fonction pour sauvegarder l'état des tickets dans un fichier JSON
def save_tickets_state():
    """Sauvegarde l'état actuel des tickets dans un fichier JSON"""
    try:
        with open('tickets_state.json', 'w') as f:
            json.dump(tickets, f, indent=4)
        app.logger.info("État des tickets sauvegardé")
    except Exception as e:
        app.logger.error(f"Erreur lors de la sauvegarde de l'état des tickets: {e}")

# Fonction pour charger l'état des tickets depuis un fichier JSON
def load_tickets_state():
    """Charge l'état des tickets depuis un fichier JSON"""
    global tickets
    try:
        if os.path.exists('tickets_state.json'):
            with open('tickets_state.json', 'r') as f:
                tickets = json.load(f)
            app.logger.info("État des tickets chargé depuis le fichier")
    except Exception as e:
        app.logger.error(f"Erreur lors du chargement de l'état des tickets: {e}")

# Fonction pour obtenir uniquement les tickets actifs (non archivés)
def get_active_tickets():
    """Retourne uniquement les tickets qui ne sont pas archivés"""
    return {tid: ticket for tid, ticket in tickets.items() if ticket.get('status') != 'archived'}

@app.route('/')
def index():
    # Tri des tickets : les tickets acceptés en haut de la liste
    active_tickets = get_active_tickets()
    sorted_tickets = sorted(active_tickets.items(), key=lambda item: (
        0 if item[1]['status'] == 'accepted' else 
        1 if item[1]['status'] == 'new' else 
        2 if item[1]['status'] == 'sent' else 3
    ))
    
    # Filtrage par client
    bnp_tickets_dict = {tid: ticket for tid, ticket in sorted_tickets if ticket.get('client') == 'Bnp Paribas Fortis'}
    infrabel_tickets_dict = {tid: ticket for tid, ticket in sorted_tickets if ticket.get('client') == 'Infrabel'}
    
    # Obtenir les statistiques pour le tableau de bord
    stats = {
        'bnp': {
            'new': len([t for _, t in bnp_tickets_dict.items() if t['status'] == 'new']),
            'accepted': len([t for _, t in bnp_tickets_dict.items() if t['status'] == 'accepted']),
            'sent': len([t for _, t in bnp_tickets_dict.items() if t['status'] == 'sent']),
            'refused': len([t for _, t in bnp_tickets_dict.items() if t['status'] == 'refused']),
            'total': len(bnp_tickets_dict)
        },
        'infrabel': {
            'new': len([t for _, t in infrabel_tickets_dict.items() if t['status'] == 'new']),
            'accepted': len([t for _, t in infrabel_tickets_dict.items() if t['status'] == 'accepted']),
            'sent': len([t for _, t in infrabel_tickets_dict.items() if t['status'] == 'sent']),
            'refused': len([t for _, t in infrabel_tickets_dict.items() if t['status'] == 'refused']),
            'total': len(infrabel_tickets_dict)
        }
    }
    
    # Onglet actif par défaut
    active_tab = request.args.get('tab', 'dashboard')
    
    # Réinitialiser le compteur de notifications quand on accède à la page principale
    global notification_count
    notification_count = 0
    
    return render_template('index.html', 
                          bnp_tickets=bnp_tickets_dict, 
                          infrabel_tickets=infrabel_tickets_dict,
                          stats=stats,
                          active_tab=active_tab,
                          notification_count=notification_count)

@app.route('/refresh')
def refresh():
    fetch_emails()
    # Retourner à l'onglet actif
    active_tab = request.args.get('tab', 'dashboard')
    return redirect(url_for('index', tab=active_tab))

@app.route('/accept/<ticket_id>')
def accept(ticket_id):
    if ticket_id not in tickets:
        flash("Ticket non trouvé.", "danger")
        return redirect(url_for('index'))
    
    try:
        # Télécharger le fichier
        file_path = download_email_attachment(ticket_id)
        if not file_path:
            flash(f"Ticket {ticket_id} : aucun fichier Excel trouvé.", "warning")
            return redirect(url_for('index'))
        
        # Vérifier si c'est un fichier .xls et le convertir en .xlsx
        if file_path.lower().endswith('.xls'):
            new_file_path = convert_xls_to_xlsx(file_path)
            if new_file_path:
                file_path = new_file_path
                flash("Fichier converti avec succès de XLS à XLSX.", "success")
            else:
                flash("Impossible de convertir le fichier XLS. Utilisation du fichier original.", "warning")
        
        # Mettre à jour le statut et le chemin du fichier
        tickets[ticket_id]['filepath'] = file_path
        tickets[ticket_id]['status'] = 'accepted'
        tickets[ticket_id]['date_traitement'] = datetime.now().strftime("%Y-%m-%d %H:%M")
        flash(f"Ticket {ticket_id} accepté et fichier prêt pour traitement.", "success")
        
        # Sauvegarder l'état après l'acceptation
        save_tickets_state()
        
    except Exception as e:
        app.logger.error(f"Erreur lors du téléchargement ou de la conversion du fichier: {e}")
        flash(f"Erreur: {e}", "danger")
    
    # Retourner à l'onglet actif
    active_tab = 'bnp' if tickets[ticket_id]['client'] == 'Bnp Paribas Fortis' else 'infrabel'
    return redirect(url_for('index', tab=active_tab))

@app.route('/refuse/<ticket_id>')
def refuse(ticket_id):
    if ticket_id in tickets:
        tickets[ticket_id]['status'] = 'refused'
        flash(f"Ticket {ticket_id} refusé.", "warning")
        # Sauvegarder l'état après le refus
        save_tickets_state()
    else:
        flash("Ticket non trouvé.", "danger")
    
    # Retourner à l'onglet actif
    active_tab = 'bnp' if tickets[ticket_id]['client'] == 'Bnp Paribas Fortis' else 'infrabel'
    return redirect(url_for('index', tab=active_tab))

@app.route('/cancel_ticket/<ticket_id>')
def cancel_ticket(ticket_id):
    """Annule un ticket accepté ou refusé pour le remettre à l'état 'new'"""
    if ticket_id not in tickets:
        flash("Ticket non trouvé.", "danger")
        return redirect(url_for('index'))
    
    ticket = tickets[ticket_id]
    
    if ticket['status'] in ['accepted', 'refused']:
        # On réinitialise le statut du ticket
        ticket['status'] = 'new'
        # On conserve le chemin du fichier si déjà téléchargé
        flash(f"Le ticket {ticket_id} a été réinitialisé. Il peut maintenant être accepté ou refusé.", "success")
        
        # Sauvegarder l'état après l'annulation
        save_tickets_state()
    else:
        flash(f"Ce ticket ne peut pas être annulé car son statut est '{ticket['status']}'.", "warning")
    
    # Retourner à l'onglet actif
    active_tab = 'bnp' if ticket['client'] == 'Bnp Paribas Fortis' else 'infrabel'
    return redirect(url_for('index', tab=active_tab))

@app.route('/ticket/<ticket_id>', methods=['GET', 'POST'])
def ticket(ticket_id):
    if ticket_id not in tickets:
        flash("Ticket non trouvé.", "danger")
        return redirect(url_for('index'))
    ticket_info = tickets[ticket_id]
    
    # Vérifier que le fichier est bien téléchargé
    if not ticket_info.get('filepath') or not os.path.exists(ticket_info['filepath']):
        flash("Le fichier associé à ce ticket n'est pas disponible.", "danger")
        return redirect(url_for('index'))
    
    filepath = ticket_info['filepath']
    
    # Si c'est encore un fichier .xls, essayer de le convertir
    if filepath.lower().endswith('.xls'):
        new_filepath = convert_xls_to_xlsx(filepath)
        if new_filepath:
            ticket_info['filepath'] = new_filepath
            tickets[ticket_id]['filepath'] = new_filepath
            filepath = new_filepath
            flash("Fichier converti en .xlsx pour le traitement.", "success")
        else:
            flash("Impossible de traiter ce fichier au format .xls. Veuillez vérifier l'installation de LibreOffice.", "danger")
            return redirect(url_for('index'))
    
    # Récupérer les valeurs existantes du fichier Excel pour les afficher dans le formulaire
    existing_values = {}
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        existing_values = {
            'sinoffcode': ws['Y3'].value or '000',
            'start_day': ws['K8'].value,
            'start_month': ws['M8'].value,
            'start_hour': ws['P8'].value,
            'start_minute': ws['R8'].value,
            'end_day': ws['T8'].value,
            'end_month': ws['N8'].value,
            'end_hour': ws['Q8'].value,
            'end_minute': ws['AA8'].value,
            'intervention_details': ws['A40'].value or ticket_info.get('intervention_details', ''),
            'explanation': ws['P40'].value or ''
        }
    except Exception as e:
        app.logger.warning(f"Impossible de récupérer les valeurs existantes du fichier: {e}")
    
    if request.method == 'POST':
        sinoffcode = request.form.get('sinoffcode', '000').zfill(3)
        start_day = request.form.get('start_day')
        start_month = request.form.get('start_month')
        start_hour = request.form.get('start_hour')
        start_minute = request.form.get('start_minute')
        end_day = request.form.get('end_day')
        end_month = request.form.get('end_month')
        end_hour = request.form.get('end_hour')
        end_minute = request.form.get('end_minute')
        intervention_details = request.form.get('intervention_details')
        explanation = request.form.get('explanation')
        current_date = request.form.get('current_date')
        
        # Tags pour le ticket
        tags = request.form.getlist('tags[]')
        ticket_info['tags'] = tags

        # Si le détail d'intervention n'existe pas encore, l'enregistrer
        if not ticket_info.get('intervention_details'):
            ticket_info['intervention_details'] = intervention_details

        try:
            # Utiliser openpyxl pour modifier le fichier .xlsx
            wb = load_workbook(filepath)
            ws = wb.active
            
            # Définir les mises à jour à effectuer
            cell_updates = {
                'Y3': sinoffcode,
                'K8': start_day,
                'M8': start_month,
                'P8': start_hour,
                'R8': start_minute,
                'T8': end_day,
                'N8': end_month,
                'Q8': end_hour,
                'AA8': end_minute,
                'B1': current_date
            }
            
            # Ajouter les cellules de texte avec condition
            if not ws['A40'].value:  # Ne pas écraser si déjà rempli
                cell_updates['A40'] = ticket_info.get('intervention_details') or intervention_details
            cell_updates['P40'] = explanation
            
            # Mettre à jour chaque cellule en gérant les cellules fusionnées
            for cell_ref, value in cell_updates.items():
                cell = ws[cell_ref]
                
                # Vérifier si la cellule est fusionnée
                if isinstance(cell, MergedCell):
                    # Trouver la cellule principale correspondante
                    for merged_range in ws.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            # Obtenir la cellule principale (en haut à gauche)
                            main_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                            main_cell.value = value
                            break
                else:
                    # Si c'est une cellule normale, la mettre à jour directement
                    cell.value = value
            
            wb.save(filepath)
            
            # Sauvegarder l'état après la mise à jour
            save_tickets_state()
            
            flash('Fichier Excel mis à jour avec succès.', 'success')
        except PermissionError:
            flash("Erreur: Le fichier est ouvert dans une autre application. Veuillez le fermer et réessayer.", "danger")
            app.logger.error(f"Erreur de permission: le fichier {filepath} est peut-être ouvert dans une autre application")
        except Exception as e:
            app.logger.error(f"Erreur lors de la mise à jour du fichier Excel: {e}")
            flash(f"Erreur lors de la mise à jour du fichier Excel: {e}", "danger")
        
        return redirect(url_for('ticket', ticket_id=ticket_id))
    
    # Passer la date du jour pour pré-remplir le champ de date
    current_date = datetime.date.today().isoformat()
    
    # Utiliser les valeurs existantes si disponibles
    return render_template('ticket.html', 
                          ticket=ticket_info, 
                          current_date=current_date,
                          values=existing_values)

@app.route('/send/<ticket_id>')
def send_ticket(ticket_id):
    if ticket_id not in tickets:
        flash('Ticket non trouvé.', "danger")
        return redirect(url_for('index'))
    ticket_info = tickets[ticket_id]
    
    if not ticket_info.get('filepath') or not os.path.exists(ticket_info['filepath']):
        flash("Le fichier associé à ce ticket n'est pas disponible.", "danger")
        return redirect(url_for('index'))
    
    try:
        # Adapter le destinataire en fonction du client
        recipients = ['Fortistech.Brussels@securitas.com']
        if ticket_info['client'] == 'Infrabel':
            recipients = ['infrabel.support@securitas.com']  # Exemple, à adapter
            
        msg = Message(subject=f"Re: {ticket_info['subject']}", recipients=recipients)
        msg.body = f"""Bonjour,

Veuillez trouver ci-joint le fichier d'intervention pour le ticket {ticket_info['ticket_id']}.

Détails de l'intervention:
{ticket_info.get('intervention_details', 'Aucun détail renseigné')}

Cordialement,
L'équipe technique Securitas Technology
"""
        with app.open_resource(ticket_info['filepath']) as fp:
            attachment_filename = os.path.basename(ticket_info['filepath'])
            # Définir le type MIME en fonction de l'extension
            if ticket_info['filepath'].lower().endswith('.xlsx'):
                mime_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            elif ticket_info['filepath'].lower().endswith('.xls'):
                mime_type = "application/vnd.ms-excel"
            else:
                mime_type = "application/octet-stream"
                
            msg.attach(attachment_filename, mime_type, fp.read())
        
        mail.send(msg)
        flash('Email envoyé avec succès.', 'success')
        
        # Marquer le ticket comme envoyé
        ticket_info['status'] = 'sent'
        ticket_info['date_envoi'] = datetime.now().strftime("%Y-%m-%d %H:%M")
        
        # Sauvegarder l'état après l'envoi
        save_tickets_state()
    except Exception as e:
        app.logger.error(f"Erreur lors de l'envoi de l'email: {e}")
        flash(f"Erreur lors de l'envoi de l'email: {e}", "danger")
    
    # Retourner à l'onglet actif
    active_tab = 'bnp' if ticket_info['client'] == 'Bnp Paribas Fortis' else 'infrabel'
    return redirect(url_for('index', tab=active_tab))

@app.route('/uploads/<path:filename>')
def download_file(filename):
    return send_from_directory(DOWNLOAD_DIR, filename)

@app.route('/get_current_time')
def get_current_time():
    now = datetime.now()
    return jsonify({
        'hour': now.hour,
        'minute': now.minute,
        'day': now.day,
        'month': now.month
    })

@app.route('/api/ticket-stats')
def ticket_stats():
    """API pour obtenir des statistiques sur les tickets pour le tableau de bord"""
    # Utiliser uniquement les tickets actifs
    active_tickets = get_active_tickets()
    
    stats = {
        'bnp': {
            'new': 0,
            'accepted': 0,
            'refused': 0,
            'sent': 0,
            'total': 0
        },
        'infrabel': {
            'new': 0,
            'accepted': 0,
            'refused': 0,
            'sent': 0,
            'total': 0
        }
    }
    
    for ticket in active_tickets.values():
        client_key = 'bnp' if ticket['client'] == 'Bnp Paribas Fortis' else 'infrabel'
        status = ticket['status']
        if status in stats[client_key]:
            stats[client_key][status] += 1
        stats[client_key]['total'] += 1
    
    return jsonify(stats)

@app.route('/search')
def search():
    """Recherche de tickets par mot-clé"""
    query = request.args.get('q', '').lower()
    if not query:
        return redirect(url_for('index'))
    
    # Utiliser uniquement les tickets actifs pour la recherche
    active_tickets = get_active_tickets()
    
    results = []
    for tid, ticket in active_tickets.items():
        if (query in ticket['subject'].lower() or 
            query in ticket['sender'].lower() or 
            query in ticket.get('intervention_details', '').lower() or
            query in ticket.get('ticket_id', '').lower()):
            results.append(ticket)
    
    return render_template('search_results.html', 
                          query=query, 
                          results=results, 
                          count=len(results))

@app.route('/add_tag/<ticket_id>', methods=['POST'])
def add_tag(ticket_id):
    """Ajouter un tag à un ticket"""
    if ticket_id not in tickets:
        return jsonify({'success': False, 'message': 'Ticket non trouvé'})
    
    tag = request.form.get('tag')
    if not tag:
        return jsonify({'success': False, 'message': 'Tag non spécifié'})
    
    if 'tags' not in tickets[ticket_id]:
        tickets[ticket_id]['tags'] = []
    
    if tag not in tickets[ticket_id]['tags']:
        tickets[ticket_id]['tags'].append(tag)
        
    # Sauvegarder l'état après l'ajout du tag
    save_tickets_state()
    
    return jsonify({
        'success': True, 
        'tags': tickets[ticket_id]['tags']
    })

@app.route('/remove_tag/<ticket_id>/<tag>')
def remove_tag(ticket_id, tag):
    """Supprimer un tag d'un ticket"""
    if ticket_id not in tickets:
        return jsonify({'success': False, 'message': 'Ticket non trouvé'})
    
    if 'tags' in tickets[ticket_id] and tag in tickets[ticket_id]['tags']:
        tickets[ticket_id]['tags'].remove(tag)
        
        # Sauvegarder l'état après la suppression du tag
        save_tickets_state()
    
    return jsonify({
        'success': True, 
        'tags': tickets[ticket_id].get('tags', [])
    })

@app.route('/filter_tickets')
def filter_tickets():
    """API pour filtrer les tickets en fonction de divers critères"""
    client = request.args.get('client', 'all')
    status = request.args.get('status', 'all')
    priority = request.args.get('priority', 'all')
    date_from = request.args.get('date_from')
    date_to = request.args.get('date_to')
    tag = request.args.get('tag')
    
    # Utiliser uniquement les tickets actifs
    active_tickets = get_active_tickets()
    
    filtered_tickets = []
    
    for tid, ticket in active_tickets.items():
        # Filtrer par client
        if client != 'all' and ticket['client'] != client:
            continue
            
        # Filtrer par statut
        if status != 'all' and ticket['status'] != status:
            continue
            
        # Filtrer par priorité
        if priority != 'all' and ticket.get('priority') != priority:
            continue
            
        # Filtrer par date
        if date_from:
            date_obj = datetime.strptime(ticket['date_reception'].split()[0], "%Y-%m-%d")
            from_obj = datetime.strptime(date_from, "%Y-%m-%d")
            if date_obj < from_obj:
                continue
                
        if date_to:
            date_obj = datetime.strptime(ticket['date_reception'].split()[0], "%Y-%m-%d")
            to_obj = datetime.strptime(date_to, "%Y-%m-%d")
            if date_obj > to_obj:
                continue
                
        # Filtrer par tag
        if tag and (not ticket.get('tags') or tag not in ticket.get('tags')):
            continue
            
        filtered_tickets.append(ticket)
    
    return jsonify({
        'tickets': filtered_tickets,
        'count': len(filtered_tickets)
    })

@app.route('/dashboard')
def dashboard():
    """Page de tableau de bord avec visualisations"""
    # Utiliser uniquement les tickets actifs pour les statistiques
    active_tickets = get_active_tickets()
    
    # Calculer les statistiques pour les graphiques
    stats = {
        'tickets_by_status': {
            'new': 0,
            'accepted': 0,
            'refused': 0,
            'sent': 0
        },
        'tickets_by_client': {
            'Bnp Paribas Fortis': 0,
            'Infrabel': 0
        },
        'tickets_by_priority': {
            'high': 0,
            'medium': 0,
            'normal': 0
        },
        'tickets_by_date': {}
    }
    
    # Date d'aujourd'hui pour calculer des statistiques sur les 30 derniers jours
    today = datetime.now().date()
    
    # Initialiser le dictionnaire de dates pour les 30 derniers jours
    for i in range(30, -1, -1):
        date = today - timedelta(days=i)
        stats['tickets_by_date'][date.strftime("%Y-%m-%d")] = 0
    
    for ticket in active_tickets.values():
        # Tickets par statut
        if ticket['status'] in stats['tickets_by_status']:
            stats['tickets_by_status'][ticket['status']] += 1
            
        # Tickets par client
        if ticket['client'] in stats['tickets_by_client']:
            stats['tickets_by_client'][ticket['client']] += 1
            
        # Tickets par priorité
        if ticket.get('priority') in stats['tickets_by_priority']:
            stats['tickets_by_priority'][ticket.get('priority')] += 1
            
        # Tickets par date
        if 'date_reception' in ticket:
            date_str = ticket['date_reception'].split()[0]  # Prendre juste la partie date
            if date_str in stats['tickets_by_date']:
                stats['tickets_by_date'][date_str] += 1
    
    return render_template('dashboard.html', stats=stats)

@app.route('/upload_logo', methods=['POST'])
def upload_logo():
    """Télécharger un nouveau logo"""
    if 'logo' not in request.files:
        flash('Aucun fichier sélectionné', 'danger')
        return redirect(url_for('index'))
        
    file = request.files['logo']
    if file.filename == '':
        flash('Aucun fichier sélectionné', 'danger')
        return redirect(url_for('index'))
        
    if file and '.' in file.filename and file.filename.rsplit('.', 1)[1].lower() in {'png', 'jpg', 'jpeg', 'gif', 'svg'}:
        filename = 'company_logo.' + file.filename.rsplit('.', 1)[1].lower()
        file.save(os.path.join('static/uploads', filename))
        flash('Logo téléchargé avec succès', 'success')
    else:
        flash('Format de fichier non autorisé', 'danger')
        
    return redirect(url_for('index'))

@app.route('/mark_all_notifications_read')
def mark_all_notifications_read():
    """Marquer toutes les notifications comme lues"""
    global notification_count
    notification_count = 0
    return jsonify({'success': True})

# Route pour les webhooks (à configurer avec votre service d'email)
@app.route('/webhook/new-email', methods=['POST'])
def email_webhook():
    """
    Point d'entrée pour les webhooks de notification d'emails.
    Permet de déclencher une vérification immédiate sans attendre le minuteur.
    """
    if request.method == 'POST':
        # Vérifier si la requête contient un token valide (à implémenter pour la sécurité)
        token = request.args.get('token', '')
        
        # Ici vous pourriez implémenter une vérification du token
        # if token != os.environ.get('WEBHOOK_SECRET_TOKEN'):
        #     return jsonify({'status': 'error', 'message': 'Invalid token'}), 403
        
        # Déclencher une vérification des emails
        try:
            fetch_emails()
            return jsonify({'status': 'success', 'message': 'Email check triggered'}), 200
        except Exception as e:
            app.logger.error(f"Erreur lors du déclenchement de la vérification des emails via webhook: {e}")
            return jsonify({'status': 'error', 'message': str(e)}), 500

@socketio.on('connect')
def handle_connect():
    """Gestion de la connexion WebSocket"""
    socketio.emit('notification_update', {'count': notification_count})
    
@socketio.on('clear_notifications')
def handle_clear_notifications():
    """Gestion de l'effacement des notifications"""
    global notification_count
    notification_count = 0
    socketio.emit('notification_update', {'count': 0})

# Gestion des erreurs 404
@app.errorhandler(404)
def page_not_found(e):
    return render_template('404.html'), 404

# Gestion des erreurs 500
@app.errorhandler(500)
def server_error(e):
    app.logger.error(f"Erreur serveur: {e}")
    return render_template('500.html'), 500

# Charger l'état des tickets au démarrage
load_tickets_state()

if __name__ == '__main__':
    # Exécuter le nettoyage au démarrage
    cleanup_old_files()
    
    # Démarrer le thread de tâches planifiées
    start_scheduled_tasks()
    
    # Démarrer le serveur
    socketio.run(app, debug=True, host='0.0.0.0', port=5000)
